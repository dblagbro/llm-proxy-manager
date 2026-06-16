"""v5.6.0 — proxy-injected read_xlsx_to_markdown tool.

Operator ask 2026-06-14: "some bots have said they can't do excel docs
— can we insert tools to do this for them?" Answered with proxy-side
tool injection: every non-streaming /v1/messages request gets a
``read_xlsx_to_markdown`` tool appended to its ``tools[]``; if the
model uses it, the proxy runs openpyxl and re-calls with the
``tool_result`` so the caller sees a final answer with file content
already incorporated.

Tests cover:
1. Registry shape + ProxyTool dataclass.
2. inject_anthropic adds the tool, is idempotent vs caller-provided
   tools, creates the list when absent.
3. find_proxy_tool_use returns None when no match + returns the tuple
   when one is present.
4. build_tool_result_message produces a valid Anthropic message.
5. Excel runner: input validation, base64 path, URL scheme rejection,
   sheet/row/col caps.
6. Source-grep contract: messages.py inject + interception are wired
   AND streaming requests do NOT inject.
"""
from __future__ import annotations

import base64
import io
from pathlib import Path
from unittest.mock import AsyncMock, patch

import pytest


# ── Registry ────────────────────────────────────────────────────────


def test_registry_contains_excel_tool():
    from app.proxy_tools import get_registry
    reg = get_registry()
    assert len(reg) >= 1
    names = {t.name for t in reg}
    assert "read_xlsx_to_markdown" in names


def test_proxy_tool_dataclass_shape():
    from app.proxy_tools import ProxyTool, get_registry
    tool = get_registry()[0]
    assert isinstance(tool, ProxyTool)
    assert isinstance(tool.anthropic_schema, dict)
    assert tool.anthropic_schema["name"] == tool.name
    assert "input_schema" in tool.anthropic_schema
    assert callable(tool.run)


# ── inject_anthropic ────────────────────────────────────────────────


def test_inject_creates_tools_list_when_absent():
    from app.proxy_tools import inject_anthropic
    body: dict = {}
    inject_anthropic(body)
    assert isinstance(body["tools"], list)
    assert any(t["name"] == "read_xlsx_to_markdown" for t in body["tools"])


def test_inject_appends_to_existing_tools():
    from app.proxy_tools import inject_anthropic
    caller_tool = {"name": "caller_thing", "input_schema": {}}
    body = {"tools": [caller_tool]}
    inject_anthropic(body)
    names = [t["name"] for t in body["tools"]]
    assert "caller_thing" in names
    assert "read_xlsx_to_markdown" in names


def test_inject_is_idempotent_against_self():
    """If the caller already supplied a tool with our name, don't
    duplicate. Prevents weird "two tools named the same" model
    behavior."""
    from app.proxy_tools import inject_anthropic
    body = {"tools": [{"name": "read_xlsx_to_markdown", "input_schema": {}}]}
    inject_anthropic(body)
    names = [t["name"] for t in body["tools"]]
    assert names.count("read_xlsx_to_markdown") == 1


# ── find_proxy_tool_use ────────────────────────────────────────────


def test_find_returns_none_when_no_tool_use():
    from app.proxy_tools import find_proxy_tool_use
    assert find_proxy_tool_use([]) is None
    assert find_proxy_tool_use([{"type": "text", "text": "hi"}]) is None


def test_find_returns_match_for_registered_tool():
    from app.proxy_tools import find_proxy_tool_use
    content = [
        {"type": "text", "text": "Let me look at that file."},
        {
            "type": "tool_use",
            "id": "toolu_abc",
            "name": "read_xlsx_to_markdown",
            "input": {"url": "https://x"},
        },
    ]
    match = find_proxy_tool_use(content)
    assert match is not None
    tool, input_obj, tid = match
    assert tool.name == "read_xlsx_to_markdown"
    assert input_obj == {"url": "https://x"}
    assert tid == "toolu_abc"


def test_find_ignores_caller_tool_use():
    from app.proxy_tools import find_proxy_tool_use
    content = [{
        "type": "tool_use",
        "id": "toolu_other",
        "name": "caller_custom_tool",
        "input": {},
    }]
    assert find_proxy_tool_use(content) is None


# ── build_tool_result_message ──────────────────────────────────────


def test_build_tool_result_message_anthropic_shape():
    from app.proxy_tools import build_tool_result_message
    msg = build_tool_result_message("toolu_abc", "# Sheet: Sales\n| a | b |")
    assert msg["role"] == "user"
    assert isinstance(msg["content"], list)
    block = msg["content"][0]
    assert block["type"] == "tool_result"
    assert block["tool_use_id"] == "toolu_abc"
    assert block["content"].startswith("# Sheet")


# ── Excel runner ───────────────────────────────────────────────────


def _make_test_xlsx_bytes() -> bytes:
    """Produce a tiny in-memory xlsx for the runner tests."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["product", "qty", "price"])
    ws.append(["widget", 10, 5.00])
    ws.append(["gizmo", 3, 12.50])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_excel_runner_renders_b64_input():
    from app.proxy_tools.excel import EXCEL_TOOL
    raw = _make_test_xlsx_bytes()
    output = await EXCEL_TOOL.run({"file_b64": base64.b64encode(raw).decode()})
    assert "## Sheet: Sales" in output
    assert "widget" in output
    assert "gizmo" in output
    # Markdown table shape
    assert "| product | qty | price |" in output
    assert "| --- | --- | --- |" in output


@pytest.mark.asyncio
async def test_excel_runner_rejects_bad_base64():
    from app.proxy_tools.excel import EXCEL_TOOL
    with pytest.raises(ValueError, match="not valid base64"):
        await EXCEL_TOOL.run({"file_b64": "not_valid_base64!!!@#$"})


@pytest.mark.asyncio
async def test_excel_runner_rejects_non_http_url():
    from app.proxy_tools.excel import EXCEL_TOOL
    with pytest.raises(ValueError, match="must use http"):
        await EXCEL_TOOL.run({"url": "file:///etc/passwd"})


@pytest.mark.asyncio
async def test_excel_runner_requires_input_source():
    from app.proxy_tools.excel import EXCEL_TOOL
    with pytest.raises(ValueError, match="file_b64 or url"):
        await EXCEL_TOOL.run({})


@pytest.mark.asyncio
async def test_excel_runner_respects_sheet_filter():
    from app.proxy_tools.excel import EXCEL_TOOL
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.create_sheet("Other")
    wb["Sheet1"]["A1"] = "in-sheet-1"
    wb["Other"]["A1"] = "in-other"
    buf = io.BytesIO(); wb.save(buf)
    output = await EXCEL_TOOL.run({
        "file_b64": base64.b64encode(buf.getvalue()).decode(),
        "sheet": "Other",
    })
    assert "in-other" in output
    assert "in-sheet-1" not in output


@pytest.mark.asyncio
async def test_excel_runner_rejects_unknown_sheet():
    from app.proxy_tools.excel import EXCEL_TOOL
    raw = _make_test_xlsx_bytes()
    output = await EXCEL_TOOL.run({
        "file_b64": base64.b64encode(raw).decode(),
        "sheet": "DoesNotExist",
    })
    assert "error" in output.lower()
    assert "DoesNotExist" in output


@pytest.mark.asyncio
async def test_excel_runner_clamps_row_cap():
    from app.proxy_tools.excel import EXCEL_TOOL
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(500):
        ws.append([f"row{i}"])
    buf = io.BytesIO(); wb.save(buf)
    output = await EXCEL_TOOL.run({
        "file_b64": base64.b64encode(buf.getvalue()).decode(),
        "max_rows": 5,
    })
    assert "row0" in output
    assert "row4" in output
    assert "row100" not in output
    assert "truncated at row cap 5" in output


# ── Wiring contracts (source-grep) ─────────────────────────────────


def test_messages_handler_injects_proxy_tools():
    """The injection MUST be wired into the request handler.

    v5.7.1 — switched the call site from sync ``inject_anthropic`` to
    async ``inject_anthropic_async`` (sources from MCP aggregator
    bridge too).
    v5.6.1 — gate lifted so streaming requests also get injection;
    server-side tool_result patcher handles the streaming round-trip."""
    src = Path("app/api/messages.py").read_text()
    assert "_proxy_tools_injected = False" in src
    idx = src.find("_proxy_tools_injected = False")
    assert idx != -1
    window = src[idx: idx + 2000]
    # Either v5.6.0 sync OR v5.7.1+ async injection must be present
    assert ("inject_anthropic(body)" in window) or ("inject_anthropic_async(body)" in window)


def test_messages_handler_intercepts_tool_use_in_response():
    """After the non-streaming response, if a proxy tool was used,
    run it + re-call. Capped at 3 hops."""
    src = Path("app/api/messages.py").read_text()
    assert "find_proxy_tool_use" in src
    assert "build_tool_result_message" in src
    # Hop cap pinned at 3
    assert "_proxy_hops < 3" in src
    # X-Proxy-Tool-Hops response header gives callers visibility
    assert 'X-Proxy-Tool-Hops' in src


def test_openpyxl_in_requirements():
    src = Path("requirements.txt").read_text()
    assert "openpyxl" in src
