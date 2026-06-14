"""v5.6.0 — read_xlsx_to_markdown proxy-injected tool.

Operator ask 2026-06-14: bots can't process Excel docs. This tool gets
appended to every non-streaming ``/v1/messages`` request; when the
model invokes it, the proxy runs openpyxl on the supplied content and
hands back a markdown rendering of every sheet so the model can
reason over the data.

Input contract:

    {
      "file_b64":      "<base64-encoded xlsx bytes>",   # one of these
      "url":           "https://...",                   #
      "sheet":         "<sheet name | optional>",       # optional
      "max_rows":      <int | default 200>,             # optional, cap
      "max_cols":      <int | default 30>,              # optional, cap
    }

Output: a markdown string with one ``## Sheet: <name>`` heading per
sheet, followed by a pipe-table of values truncated to the row/col
caps. Errors come back as the string ``error: <reason>`` so the model
sees them inline.

Size caps:
- file payload limited to 5 MB to keep prompt-token cost bounded.
- max_rows/max_cols default low; the model can re-call with larger
  bounds if it actually needs more data.

Security:
- ``url`` mode fetches only ``http://`` / ``https://`` (no ``file://``,
  no SSRF defence beyond scheme check today — operator-set fleet uses
  trusted bot infrastructure; v5.6.x can add allow-list if needed).
- No openpyxl macro execution: ``load_workbook(data_only=True)``
  bypasses formulas; cell values come from the cached eval result.
"""
from __future__ import annotations

import base64
import io
from typing import Any, Dict

from app.proxy_tools import ProxyTool


_MAX_BYTES = 5 * 1024 * 1024  # 5 MB
_DEFAULT_MAX_ROWS = 200
_DEFAULT_MAX_COLS = 30
_TOOL_NAME = "read_xlsx_to_markdown"


_ANTHROPIC_SCHEMA: dict = {
    "name": _TOOL_NAME,
    "description": (
        "Read an Excel (.xlsx) spreadsheet and return its contents as "
        "markdown tables. Supply EITHER a base64-encoded file blob "
        "(file_b64) OR a https URL the proxy can fetch. Each sheet "
        "becomes a `## Sheet: <name>` section with a pipe-table of cell "
        "values. Defaults to first 200 rows × 30 columns; pass max_rows "
        "/ max_cols to widen. Use this when the user asks about a "
        "spreadsheet — do not say you can't read xlsx files, just call "
        "this tool."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "file_b64": {
                "type": "string",
                "description": (
                    "Base64-encoded xlsx file bytes. Max 5 MB decoded. "
                    "Use this when you have the file contents inline."
                ),
            },
            "url": {
                "type": "string",
                "description": (
                    "https URL of the xlsx file. Proxy fetches it "
                    "server-side. Use when the file lives at a known "
                    "URL the proxy can reach."
                ),
            },
            "sheet": {
                "type": "string",
                "description": (
                    "If set, render only this sheet. Otherwise all sheets."
                ),
            },
            "max_rows": {
                "type": "integer",
                "description": "Row cap per sheet (default 200).",
            },
            "max_cols": {
                "type": "integer",
                "description": "Column cap per sheet (default 30).",
            },
        },
    },
}


_OPENAI_SCHEMA: dict = {
    "type": "function",
    "function": {
        "name": _TOOL_NAME,
        "description": _ANTHROPIC_SCHEMA["description"],
        "parameters": _ANTHROPIC_SCHEMA["input_schema"],
    },
}


async def _fetch_bytes(input_obj: Dict[str, Any]) -> bytes:
    """Resolve the input into bytes. Raises ValueError on bad input."""
    if isinstance(input_obj.get("file_b64"), str) and input_obj["file_b64"]:
        try:
            raw = base64.b64decode(input_obj["file_b64"], validate=True)
        except Exception as exc:
            raise ValueError(f"file_b64 is not valid base64: {exc}")
        if len(raw) > _MAX_BYTES:
            raise ValueError(
                f"file_b64 too large: {len(raw)} bytes > {_MAX_BYTES} cap"
            )
        return raw

    url = input_obj.get("url")
    if isinstance(url, str) and url:
        if not (url.startswith("https://") or url.startswith("http://")):
            raise ValueError(
                "url must use http:// or https://; got "
                f"{url.split(':', 1)[0]}://"
            )
        import httpx
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            resp = await client.get(url)
            resp.raise_for_status()
        if len(resp.content) > _MAX_BYTES:
            raise ValueError(
                f"url body too large: {len(resp.content)} bytes > {_MAX_BYTES} cap"
            )
        return resp.content

    raise ValueError("must supply either file_b64 or url")


def _render_workbook(
    data: bytes,
    sheet_filter: str | None,
    max_rows: int,
    max_cols: int,
) -> str:
    """Render bytes as markdown. Pure function — no I/O."""
    import openpyxl
    wb = openpyxl.load_workbook(
        io.BytesIO(data), data_only=True, read_only=True,
    )
    out: list[str] = []
    sheet_names = wb.sheetnames
    if sheet_filter:
        if sheet_filter not in sheet_names:
            return (
                f"error: sheet {sheet_filter!r} not found; "
                f"available: {sheet_names}"
            )
        sheet_names = [sheet_filter]

    for sname in sheet_names:
        ws = wb[sname]
        out.append(f"## Sheet: {sname}")
        rows_yielded = 0
        body_rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            if rows_yielded >= max_rows:
                out.append(
                    f"_(truncated at row cap {max_rows}; "
                    f"sheet has more rows)_"
                )
                break
            trimmed = list(row)[:max_cols]
            body_rows.append([
                "" if v is None else str(v).replace("|", r"\|")
                for v in trimmed
            ])
            rows_yielded += 1
        if body_rows:
            ncols = max(len(r) for r in body_rows)
            header = body_rows[0] if rows_yielded else []
            # Pad first row out to ncols
            while len(header) < ncols:
                header.append("")
            out.append("| " + " | ".join(header) + " |")
            out.append("|" + " --- |" * ncols)
            for r in body_rows[1:]:
                while len(r) < ncols:
                    r.append("")
                out.append("| " + " | ".join(r) + " |")
        else:
            out.append("_(empty)_")
        out.append("")
    return "\n".join(out).strip()


async def _run_excel(input_obj: dict) -> str:
    """Coroutine entrypoint that ProxyTool.run wraps."""
    data = await _fetch_bytes(input_obj)
    sheet = input_obj.get("sheet")
    max_rows = int(input_obj.get("max_rows") or _DEFAULT_MAX_ROWS)
    max_cols = int(input_obj.get("max_cols") or _DEFAULT_MAX_COLS)
    if max_rows <= 0 or max_rows > 5000:
        raise ValueError("max_rows must be in (0, 5000]")
    if max_cols <= 0 or max_cols > 200:
        raise ValueError("max_cols must be in (0, 200]")
    return _render_workbook(data, sheet, max_rows, max_cols)


EXCEL_TOOL = ProxyTool(
    name=_TOOL_NAME,
    anthropic_schema=_ANTHROPIC_SCHEMA,
    openai_schema=_OPENAI_SCHEMA,
    run=_run_excel,
)
